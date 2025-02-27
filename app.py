from flask import Flask, render_template, request, send_file
import pdfplumber
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import gc

def extraer_datos(pdf_path):
    datos = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()
            if texto:
                lineas = texto.split('\n')
                numero_albarán = ""
                for i, linea in enumerate(lineas):
                    if "Decl. goods it Nr." in linea:
                        if i + 1 < len(lineas):
                            partes = lineas[i + 1].strip().split()
                            if len(partes) >= 6:
                                numero_partida = partes[0]
                                numero_bultos = partes[4]
                                if len(partes[5]) == 10 and partes[5].isdigit():
                                    descripcion = " ".join(partes[6:])
                                else:
                                    descripcion = " ".join(partes[5:])
                    if "UCR [12 08] Gross mass [18 04]" in linea:
                        if i + 1 < len(lineas):
                            partes_peso = lineas[i + 1].strip().split()
                            if len(partes_peso) >= 2:
                                numero_albarán = partes_peso[0]
                                peso = float(partes_peso[1])
                                if peso > 0:
                                    datos.append([numero_partida, int(numero_bultos), descripcion, numero_albarán, peso])
    return datos

def estilizar_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    
    # Estilos de color
    fill_bultos = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    fill_alternate_1 = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    fill_alternate_2 = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        if row[1].value > 1:
            for cell in row:
                cell.fill = fill_bultos
        else:
            fill = fill_alternate_1 if i % 2 == 0 else fill_alternate_2
            for cell in row:
                cell.fill = fill
    
    # Centrar la columna de número de bultos
    for cell in ws["B"]:
        cell.alignment = Alignment(horizontal="center")
    
    wb.save(excel_path)
    del wb, ws, fill_bultos, fill_alternate_1, fill_alternate_2
    gc.collect()  # Forzar recolección de basura

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return "No file part"
    file = request.files['file']
    if file.filename == '':
        return "No selected file"
    
    try:
        pdf_path = "uploaded.pdf"
        file.save(pdf_path)
    except Exception as e:
        return f"Error saving PDF: {e}"
    
    try:
        datos = extraer_datos(pdf_path)
    except Exception as e:
        return f"Error extracting data from PDF: {e}"
    
    try:
        df = pd.DataFrame(datos, columns=["Número de partida", "Número de bultos", "Descripción de la mercancía", "Número de albarán", "Peso"])
        excel_path = "output.xlsx"
        df.to_excel(excel_path, index=False)
        del df  # Liberar memoria
    except Exception as e:
        return f"Error creating Excel file: {e}"
    
    try:
        estilizar_excel(excel_path)
    except Exception as e:
        return f"Error styling Excel file: {e}"
    
    try:
        return send_file(excel_path, as_attachment=True)
    except Exception as e:
        return f"Error sending Excel file: {e}"

if __name__ == '__main__':
    app.run(debug=True)
